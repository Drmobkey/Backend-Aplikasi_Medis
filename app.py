from flask import Flask, render_template, request, jsonify
import numpy as np
import cv2
import pydicom
import tensorflow as tf
import base64
from io import BytesIO
from PIL import Image
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
import io

app = Flask(__name__)
CORS(app)

# Pastikan model tersedia atau ganti path sesuai lokasi Anda
try:
    model = tf.keras.models.load_model("best_model.keras")
except:
    print("Model file not found. Please check the path.")
    # Placeholder dummy model logic if file missing (for testing structure)
    model = None 

classes = ['ACCEPT', 'ARTEFAK', 'EXPOSURE', 'POSITION']

# Data tindak lanjut
tindaklanjut = [
    {
        "dataid": "POSITION",
        "pengulangan": [
            "Mengalami rotasi/ posisi miring",
            "Pasien salah dalam orientasi"
        ],
        "diskripsi": [
            "Posisi anatomi yang tidak tepat",
            "Posisi tidak true, mengalami rotasi baik ke internal maupun eksternal"
        ],
        "tindaklanjut": [
            "Memastikan semua radiografer memiliki sertifikasi kompetensi",
            "Adopsi protokol standar posisi pasien dan teknik citra",
            "Menyediakan panduan prosedur tertulis",
            "Membangun komunikasi efektif dengan pasien",
            "Audit rutin citra radiografi untuk pola kesalahan"
        ]
    },
    {
        "dataid": "ARTEFAK",
        "pengulangan": [
            "Obyek yang dikenali",
            "Grid line atau artefak sejenis",
            "Ketidakseragaman atau defect/cacat terlihat"
        ],
        "diskripsi": [
            "Obyek seperti kancing, perhiasan, shield",
            "Artefak interferensi elektromagnetik",
            "Artefak detector seperti piksel mati"
        ],
        "tindaklanjut": [
            "Mengurangi gangguan elektromagnetik",
            "Memastikan pasien bebas dari bahan logam",
            "Pemeliharaan dan kalibrasi alat rutin",
            "Screening pasien sebelum pemeriksaan",
            "Audit rutin citra radiografi"
        ]
    },
    {
        "dataid": "EXPOSURE",
        "pengulangan": [
            "Under eksposure",
            "Over eksposure/saturasi"
        ],
        "diskripsi": [
            "Noise berlebih akibat eksposure kurang",
            "Over/under density akibat faktor teknik"
        ],
        "tindaklanjut": [
            "Adopsi protokol standar",
            "Pencatatan hasil & koreksi",
            "Komunikasi efektif antara radiografer-teknisi",
            "Pemeliharaan rutin alat",
            "Audit citra radiografi berkala"
        ]
    },
    {
        "dataid": "ACCEPT",
        "pengulangan": [],
        "diskripsi": [],
        "tindaklanjut": []
    },
]

def read_dicom_as_rgb(dicom_bytes):
    dicom = pydicom.dcmread(BytesIO(dicom_bytes))
    image = dicom.pixel_array
    # Normalisasi jika perlu agar rentang 0-255 uint8
    if image.max() > 255:
        image = (image / image.max()) * 255
    image = np.uint8(image)
    
    if len(image.shape) == 2:
        image = cv2.cvtColor(image, cv2.COLOR_GRAY2RGB)
    return image

def preprocess_image(image):
    image_resized = cv2.resize(image, (256, 256))
    return image_resized / 255.0

def encode_image_to_base64(image):
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    pil_image = Image.fromarray(np.uint8(image_rgb))
    buffer = BytesIO()
    pil_image.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode()

@app.route('/')
def index():
    return render_template("index.html")

@app.route('/predict', methods=['POST'])
def predict():
    if model is None:
        return jsonify({"error": "Model not loaded"}), 500
        
    file = request.files['file']
    dicom_bytes = file.read()

    image = read_dicom_as_rgb(dicom_bytes)
    processed = preprocess_image(image)
    input_tensor = np.expand_dims(processed, axis=0)

    preds = model.predict(input_tensor)[0]
    predicted_index = np.argmax(preds)
    predicted_class = classes[predicted_index]

    base64_image = encode_image_to_base64(image)

    probs = sorted(
        [{"class_name": classes[i], "percent": f"{preds[i]*100:.2f}"} for i in range(len(classes))],
        key=lambda x: float(x["percent"]),
        reverse=True
    )

    matched_action = next((item for item in tindaklanjut if item["dataid"] == predicted_class), None)

    return jsonify({
        "predicted_class": predicted_class,
        "probabilities": probs,
        "image": base64_image,
        "action": matched_action
    })

@app.route('/predict-multiple', methods=['POST'])
def predict_multiple():
    if model is None:
        return jsonify({"error": "Model not loaded"}), 500

    files = request.files.getlist('files')
    if not files:
        return jsonify({"error": "No files uploaded."}), 400

    result_counter = {cls: 0 for cls in classes}
    excel_rows = [] 
    error_files = []

    # 1. Proses Prediksi
    for idx, file in enumerate(files, 1):
        filename = file.filename
        dicom_bytes = file.read()
        try:
            # Proses file DICOM
            image = read_dicom_as_rgb(dicom_bytes)
            processed = preprocess_image(image)
            input_tensor = np.expand_dims(processed, axis=0)

            preds = model.predict(input_tensor)[0]
            predicted_index = np.argmax(preds)
            predicted_class = classes[predicted_index]
            confidence_percent = preds[predicted_index] * 100

            result_counter[predicted_class] += 1
            
            # Simpan data baris: No, Nama File, Kelas, Persentase
            excel_rows.append([idx, filename, predicted_class, f"{confidence_percent:.2f}%"])
            
        except Exception as e:
            print(f"Failed to process {filename}: {e}")
            error_files.append(filename)
            excel_rows.append([idx, filename, "ERROR", "0%"])

    # 2. Membuat Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Klasifikasi"

    # --- Definition Styles ---
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    center_aligned = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # -- Bagian A: Tabel Utama --
    headers = ["No", "Nama File", "Prediksi Kelas Tertinggi", "Persentase"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border

    for row_data in excel_rows:
        ws.append(row_data)
        current_row = ws.max_row
        for col_idx in range(1, 5): 
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            if col_idx == 1: 
                cell.alignment = center_aligned

    last_data_row = len(excel_rows) + 1
    
    # -- Bagian B: Tabel Ringkasan untuk Chart --
    summary_start_row = last_data_row + 4
    
    title_cell = ws.cell(row=summary_start_row, column=1, value="Ringkasan Kelas")
    title_cell.font = Font(bold=True, size=12)

    header_row_summary = summary_start_row + 1
    ws.cell(row=header_row_summary, column=1, value="Kelas")
    ws.cell(row=header_row_summary, column=2, value="Jumlah")

    for col_idx in range(1, 3):
        cell = ws.cell(row=header_row_summary, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
        cell.border = thin_border

    current_row = summary_start_row + 2
    for cls in classes:
        count = result_counter[cls]
        c1 = ws.cell(row=current_row, column=1, value=cls)   
        c2 = ws.cell(row=current_row, column=2, value=count) 
        
        c1.border = thin_border
        c2.border = thin_border
        c2.alignment = center_aligned 

        current_row += 1

    summary_end_row = current_row - 1

    # -- Bagian C: Auto Adjust Column Width --
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # -- Bagian D: Membuat Bar Chart --
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Distribusi Hasil Klasifikasi"
    chart.y_axis.title = 'Jumlah Deteksi' # <--- Diganti sesuai request
    chart.x_axis.title = 'Kelas'
    
    # Konfigurasi Label Data
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True       # Tampilkan Nilai (Angka)
    chart.dataLabels.showCatName = True   # Tampilkan Nama Kategori (Kelas)
    chart.dataLabels.showSerName = False  # <--- HILANGKAN Series Name ("Jumlah")
    
    chart.legend.position = 'b'

    data = Reference(ws, min_col=2, min_row=summary_start_row+1, max_row=summary_end_row)
    cats = Reference(ws, min_col=1, min_row=summary_start_row+2, max_row=summary_end_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    
    chart.width = 15 
    chart.height = 10 

    chart_position = f"A{summary_end_row + 3}"
    ws.add_chart(chart, chart_position)

    # 3. Simpan Excel
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    excel_base64 = base64.b64encode(excel_buffer.getvalue()).decode()

    # 4. JSON Summary
    total_files = len(files)
    result_summary = []
    for cls in classes:
        count = result_counter[cls]
        percent = (count / total_files) * 100 if total_files else 0
        result_summary.append({
            "class_name": cls,
            "count": count,
            "percent": f"{percent:.2f}", 
        })
    result_summary.sort(key=lambda x: x["count"], reverse=True)
    
    top_class_overall = result_summary[0]["class_name"] if result_summary else "ACCEPT"
    matched_action = next((item for item in tindaklanjut if item["dataid"] == top_class_overall), None)

    response_data = {
        "total_files": total_files,
        "summary": result_summary,
        "excel_base64": excel_base64,
        "action": matched_action
    }

    if error_files:
        response_data["errors"] = {
            "count": len(error_files),
            "files": error_files
        }

    return jsonify(response_data)


if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5001, debug=True)